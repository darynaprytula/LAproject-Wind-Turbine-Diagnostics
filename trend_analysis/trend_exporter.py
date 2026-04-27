import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

FILL_WARNING = PatternFill("solid", start_color="FFB347")
FILL_ALARM   = PatternFill("solid", start_color="FF4C4C")
FILL_OK      = PatternFill("solid", start_color="C6EFCE")

HDR_FILL = PatternFill("solid", start_color="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DAT_FONT = Font(name="Arial", size=10)
CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left",   vertical="center")


def _style_sheet(ws, status_col_idx: int):
    """Apply header styles and color Status column rows."""

    for cell in ws[1]:
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = CTR

    for row in ws.iter_rows(min_row=2):
        status_cell = row[status_col_idx - 1]
        status_val  = str(status_cell.value).upper() if status_cell.value else ""

        if status_val == "ALARM":
            fill = FILL_ALARM
        elif status_val == "WARNING":
            fill = FILL_WARNING
        else:
            fill = FILL_OK

        for cell in row:
            cell.font      = DAT_FONT
            cell.alignment = LEFT

        status_cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def save_trend_excel(
    metrics_df: pd.DataFrame,
    fsc_df: pd.DataFrame,
    output_path: str
) -> None:
    """
    Save trend analysis results to Excel with coloured Status column.

    Metrics sheet — one sheet per turbine.
    FSC sheet     — one sheet per turbine.

    WARNING rows  → orange Status cell
    ALARM rows    → red    Status cell
    OK rows       → green  Status cell

    Parameters:
    metrics_df  : dataframe returned by metrics_trend_analysis()
    fsc_df      : dataframe returned by fsc_trend_analysis()
    output_path : output .xlsx file path
    """

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        for turbine, grp in metrics_df.groupby("Turbine"):
            sheet_name = f"M_{turbine}"[:31]
            grp.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            status_col = grp.columns.get_loc("Status") + 1
            _style_sheet(ws, status_col)

        for turbine, grp in fsc_df.groupby("Turbine"):
            sheet_name = f"FSC_{turbine}"[:31]
            grp.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            status_col = grp.columns.get_loc("Status") + 1
            _style_sheet(ws, status_col)

    print(f"Trend Excel saved: {output_path}")
