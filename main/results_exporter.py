import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNS = [
    ("start_ms",             "Signal start time (ms)"),
    ("end_ms",               "Signal end time (ms)"),
    ("sensor_num",           "Sensor №"),
    ("duration",             "Duration (s)"),
    ("operating_conditions", "Operating conditions"),
    ("signal_type",          "Signal type"),
    ("resampling_mode",      "Resampling"),
    ("frequency_span",       "Frequency span"),
    ("lor",                  "Lines of resolution"),

    ("MaxPeak",              "Max peak"),
    ("PeakToPeak",           "Peak-to-peak"),
    ("RMS",                  "RMS"),
    ("CrestFactor",          "Crest factor"),
    ("KFactor",              "K-factor"),
    ("ImpulseFactor",        "Impulse factor"),
    ("Skewness",             "Skewness"),
    ("Kurtosis",             "Kurtosis"),

    ("WindSpeed_mean",       "Wind speed mean (m/s)"),
    ("Power_mean",           "Power mean (kW)"),
    ("RPM_mean",             "RPM mean"),
]


FSC_COMBOS = [
    ("Rotor_RBG", "BPFO"), ("Rotor_RBG", "BPFI"), ("Rotor_RBG", "FTF"), ("Rotor_RBG", "BSF2"),
    ("GBX_In_BRG1", "BPFO"), ("GBX_In_BRG1", "BPFI"), ("GBX_In_BRG1", "FTF"),
    ("GBX_In_BRG1", "BSF2"), ("GBX_OUT_BRG2", "BPFO"), ("GBX_OUT_BRG2", "BPFI"),
    ("GBX_OUT_BRG2", "FTF"), ("GBX_OUT_BRG2", "BSF2"), ("GBX_OUT_BRG1", "BPFO"),
    ("GBX_OUT_BRG1", "BPFI"), ("GBX_OUT_BRG1", "FTF"), ("GBX_OUT_BRG1", "BSF2"),
    ("Gen_DE_BRG", "BPFO"), ("Gen_DE_BRG", "BPFI"), ("Gen_DE_BRG", "FTF"), ("Gen_DE_BRG", "BSF2"),
    ("Gen_NDE_RBG", "BPFO"), ("Gen_NDE_RBG", "BPFI"), ("Gen_NDE_RBG", "FTF"),
    ("Gen_NDE_RBG", "BSF2"), ("GBX-Out_Pinion", "GMF1x"), ("GBX-Out_Pinion", "GMF2x"),
    ("GBX-Out_Wheel", "GMF1x"), ("GBX-Out_Wheel", "GMF2x"),
    ("GBX-Mid_Sun", "GMF1x"), ("GBX-Mid_Sun", "GMF2x"), ("GBX-Mid_Sun", "GMF3x"),
]


INT_COLS  = {"start_ms", "end_ms"}

VIBR_COLS = {
    "MaxPeak",
    "PeakToPeak",
    "RMS",
    "CrestFactor",
    "KFactor",
    "ImpulseFactor",
    "Skewness",
    "Kurtosis",
}

OPER_COLS = {"WindSpeed_mean", "Power_mean", "RPM_mean"}

HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DAT_FONT  = Font(name="Arial", size=10)

CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

VIBR_FILL = PatternFill("solid", start_color="E2EFDA")
OPER_FILL = PatternFill("solid", start_color="DDEBF7")
FSC_FILL  = PatternFill("solid", start_color="FFF2CC")


COL_WIDTHS_BASE = [
    18, 18, 10, 12, 22, 12, 12, 14, 18,
    12, 14, 12, 14, 12, 12, 22, 18, 12,
]


def load_fsc(fsc_csv_path: str) -> dict:
    """
    Load FSC results from CSV file.

    parameters:
    fsc_csv_path - path to FSC CSV file

    returns:
    dictionary:
        (turbine, filename, component, characteristic) -> (frequency, amplitude)
    """

    fsc = {}

    try:
        with open(fsc_csv_path, "r", encoding="utf-8") as f:
            for row in csv.DictReader(f):

                key = (
                    row["turbine"],
                    row["name"],
                    row["component"],
                    row["characteristic"],
                )

                try:
                    freq = float(row["target_frequency_hz"]) if row["target_frequency_hz"] else None
                    amp  = float(row["amplitude"]) if row["amplitude"] else None
                except ValueError:
                    freq, amp = None, None

                fsc[key] = (freq, amp)

    except FileNotFoundError:
        print(f"[WARNING] FSC file not found: {fsc_csv_path}")

    return fsc


def export(grouped_rows: dict, output_path: str, fsc_csv_path: str = "fsc_results.csv") -> None:
    """
    Export processed data into Excel with FSC features.

    parameters:
    grouped_rows - dict: turbine -> list of rows
    output_path - output Excel file path
    fsc_csv_path - path to FSC CSV file
    """

    fsc = load_fsc(fsc_csv_path)

    base_keys   = [c[0] for c in COLUMNS]
    base_titles = [c[1] for c in COLUMNS]

    fsc_titles = []
    for comp, char in FSC_COMBOS:
        label = f"{comp}_{char}"
        fsc_titles.append(f"{label} Freq (Hz)")
        fsc_titles.append(f"{label} Amp")

    all_titles = base_titles + fsc_titles

    wb = Workbook()
    wb.remove(wb.active)

    for turbine in sorted(grouped_rows.keys()):
        rows = grouped_rows[turbine]
        ws = wb.create_sheet(title=turbine[:31])

        ws.append(all_titles)

        for col in range(1, len(all_titles) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CTR

        for r_idx, row in enumerate(rows, start=2):

            for c_idx, key in enumerate(base_keys, start=1):
                val = row.get(key)
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.font = DAT_FONT
                cell.alignment = LEFT

                if key in VIBR_COLS:
                    cell.fill = VIBR_FILL
                elif key in OPER_COLS:
                    cell.fill = OPER_FILL

            fname = row.get("filename") or ""

            for i, (comp, char) in enumerate(FSC_COMBOS):
                freq_col = len(base_keys) + i * 2 + 1
                amp_col  = freq_col + 1

                key = (turbine, fname, comp, char)
                freq, amp = fsc.get(key, (None, None))

                ws.cell(row=r_idx, column=freq_col, value=freq).fill = FSC_FILL
                ws.cell(row=r_idx, column=amp_col,  value=amp).fill  = FSC_FILL

        print(f"[{turbine}] -> {len(rows)} rows")

    wb.save(output_path)
    print(f"\nSaved: {output_path}")
